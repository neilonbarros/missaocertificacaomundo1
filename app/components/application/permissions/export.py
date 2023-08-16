import csv
from io import BytesIO

from django.http import HttpRequest, HttpResponse
from openpyxl import Workbook

from app import decorators as appdecorators
from app import models as appmodels
from app import packages as apppackages


@appdecorators.authenticated.is_authenticated()
@appdecorators.permissions.validate(
    file=__file__,
)
def mycsv(
    request: HttpRequest,
    session_user: apppackages.utils.Session,
):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="permissions.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "permission",
            "jobposition",
            "department",
        ]
    )

    result = appmodels.ApplicationPermissions.objects.all().select_related(
        "jobposition",
    ).values_list(
        "permission",
        "jobposition__jobposition",
        "jobposition__department__department",
    )
    for row in result:
        writer.writerow(row)

    return response


@appdecorators.authenticated.is_authenticated()
@appdecorators.permissions.validate(
    file=__file__,
)
def myxlsx(
    request: HttpRequest,
    session_user: apppackages.utils.Session,
):
    excelfile = BytesIO()
    workbook = Workbook()
    worksheet = workbook.create_sheet(title="export", index=0)

    result = appmodels.ApplicationPermissions.objects.all().select_related(
        "jobposition",
    ).values_list(
        "permission",
        "jobposition__jobposition",
        "jobposition__department__department",
    )

    columns: list[str] = [
            "permission",
            "jobposition",
            "department",
        ]
    row_index: int = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_index, column=col_num)
        cell.value = column_title

    # Iterate through all coins
    for _, row in enumerate(result, 1):
        row_index += 1

        # Define the data for each cell in the row
        row_data = [
            row[0],
            row[1],
            row[2],
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_index, column=col_num)
            cell.value = cell_value
    workbook.save(excelfile)

    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment; filename="permissions.xlsx"'
    response.write(excelfile.getvalue())

    return response
