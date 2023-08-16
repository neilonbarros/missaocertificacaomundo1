import csv
import os
from io import BytesIO

from django.http import HttpRequest, HttpResponse
from openpyxl import Workbook

from app import decorators as appdecorators
from app import models as appmodels
from app import packages as apppackages


@appdecorators.authenticated.is_authenticated()
@appdecorators.permissions.validate(
    file=__file__,
)
def mycsv(
    request: HttpRequest,
    session_user: apppackages.utils.Session,
):
    response = HttpResponse(content_type="text/csv")
    response[
        "Content-Disposition"
    ] = f'attachment; filename="{os.path.basename(os.path.dirname(__file__))}.csv"'  # noqa

    writer = csv.writer(response)
    writer.writerow(
        [
            "id",
            "status",
            "fullname",
            "department",
            "jobposition",
        ]
    )

    result = (
        appmodels.ApplicationPeoples.objects.all()
        .select_related(
            "department",
            "jobposition",
        )
        .values_list(
            "id",
            "status",
            "fullname",
            "department__department",
            "jobposition__jobposition",
        )
    )
    for row in result:
        writer.writerow(row)

    return response


@appdecorators.authenticated.is_authenticated()
@appdecorators.permissions.validate(
    file=__file__,
)
def myxlsx(
    request: HttpRequest,
    session_user: apppackages.utils.Session,
):
    excelfile = BytesIO()
    workbook = Workbook()
    worksheet = workbook.create_sheet(title="export", index=0)

    result = (
        appmodels.ApplicationPeoples.objects.all()
        .select_related(
            "department",
            "jobposition",
        )
        .values_list(
            "id",
            "status",
            "fullname",
            "department__department",
            "jobposition__jobposition",
        )
    )

    columns: list[str] = [
        "id",
        "status",
        "fullname",
        "department",
        "jobposition",
    ]
    row_index: int = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_index, column=col_num)
        cell.value = column_title

    # Iterate through all coins
    for _, row in enumerate(result, 1):
        row_index += 1

        # Define the data for each cell in the row
        row_data = [
            row[0],
            row[1],
            row[2],
            row[3],
            row[4],
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_index, column=col_num)
            cell.value = cell_value
    workbook.save(excelfile)

    response = HttpResponse(content_type="application/vnd.ms-excel")
    response[
        "Content-Disposition"
    ] = f'attachment; filename="{os.path.basename(os.path.dirname(__file__))}.xlsx"'  # noqa
    response.write(excelfile.getvalue())

    return response
